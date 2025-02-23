import gradio as gr
import os
import io
import pdfplumber
import docx
from openpyxl import load_workbook
import shutil
import threading
from datetime import datetime
from typing import List, Dict, Any, Generator
from session_manager import SessionManager
from huggingface_hub import InferenceClient

# Initialize session manager and get HF API key
session_manager = SessionManager()
HF_API_KEY = os.getenv("HF_API_KEY")

# Create uploads directory if it doesn't exist
os.makedirs("uploads", exist_ok=True)

# Model endpoints configuration
MODEL_ENDPOINTS = {
    "Qwen2.5-72B-Instruct": "https://api-inference.huggingface.co/models/Qwen/Qwen2.5-72B-Instruct",
    "Llama3.3-70B-Instruct": "https://api-inference.huggingface.co/models/meta-llama/Llama-3.3-70B-Instruct",
    "Qwen2.5-Coder-32B-Instruct": "https://api-inference.huggingface.co/models/Qwen/Qwen2.5-Coder-32B-Instruct",
}

def query_model(model_name: str, messages: List[Dict[str, str]]) -> Generator[str, None, None]:
    """Query a single model with the chat history and stream the response"""
    endpoint = MODEL_ENDPOINTS[model_name]
    
    # Build full conversation history for context
    conversation = "\n".join([f"{msg['role']}: {msg['content']}" for msg in messages])
    
    # System prompt configuration
    system_prompts = {
        "Qwen2.5-72B-Instruct": "Collaborate with other experts. Previous discussion:\n{conversation}",
        "Llama3.3-70B-Instruct": (
            "<|begin_of_text|><|start_header_id|>system<|end_header_id|>\n"
            f"Build upon this discussion:\n{conversation}<|eot_id|>\n"
            "<|start_header_id|>assistant<|end_header_id|>\nMy contribution:"
        ),
        "Qwen2.5-Coder-32B-Instruct": (
            f"<|im_start|>system\nTechnical discussion context:\n{conversation}<|im_end|>\n"
            "<|im_start|>assistant\nTechnical perspective:"
        )
    }

    client = InferenceClient(base_url=endpoint, token=HF_API_KEY)

    try:
        messages = [
            {"role": "system", "content": system_prompts[model_name].format(conversation=conversation)},
            {"role": "user", "content": "Continue the expert discussion"}
        ]
        
        stream = client.chat.completions.create(
            messages=messages,
            stream=True,
            max_tokens=2048,
            temperature=0.5,
            top_p=0.7
        )

        for chunk in stream:
            content = chunk.choices[0].delta.content or ""
            yield content

    except Exception as e:
        yield f"{model_name} error: {str(e)}"

def respond(message: str, history: List[List[str]], session_id: str) -> Generator[str, None, None]:
    """Handle sequential model responses with context preservation and streaming"""
    # Load or initialize session

    session = session_manager.load_session(session_id)
    if not isinstance(session, dict) or "history" not in session:
        session = {"history": []}

    # Build context from session history
    messages = []
    for entry in session["history"]:
        if entry["type"] == "user":
            messages.append({"role": "user", "content": entry["content"]})
        else:
            messages.append({"role": "assistant", "content": f"{entry['model']}: {entry['content']}"})

    # Add current message
    messages.append({"role": "user", "content": message})

    # Add file content to message


    session["history"].append({
        "timestamp": datetime.now().isoformat(),
        "type": "user",
        "content": message
    })

    # Model responses
    model_names = ["Qwen2.5-Coder-32B-Instruct", "Qwen2.5-72B-Instruct", "Llama3.3-70B-Instruct"]
    model_colors = ["🔵", "🟣", "🟡"]
    responses = {}

    # Initialize responses
    for model_name in model_names:
        responses[model_name] = ""

    # Stream responses from each model
    for i, model_name in enumerate(model_names):
        yield f"{model_colors[i]} {model_name} is thinking..."
        
        full_response = ""
        for chunk in query_model(model_name, messages):
            full_response += chunk
            yield f"{model_colors[i]} **{model_name}**\n{full_response}"

        # Update session history and messages
        session["history"].append({
            "timestamp": datetime.now().isoformat(),
            "type": "assistant",
            "model": model_name,
            "content": full_response
        })
        messages.append({"role": "assistant", "content": f"{model_name}: {full_response}"})
        responses[model_name] = full_response

    # Save final session state
    session_manager.save_session(session_id, session)

    # Return final combined response (optional)
    combined_response = ""
    for i, model_name in enumerate(model_names):
        combined_response += f"{model_colors[i]} **{model_name}**\n{responses[model_name]}\n\n"
    yield combined_response

# Create the Gradio interface
with gr.Blocks() as demo:
    gr.Markdown("## Multi-LLM Collaboration Chat")

    with gr.Row():
        session_id = gr.State(session_manager.create_session)
        new_session = gr.Button("🔄 New Session")

    chatbot = gr.Chatbot(height=600)
    save_history = gr.Checkbox(label="Save Conversation History", value=True)

    def on_new_session():
        new_id = session_manager.create_session()
        return new_id, []

    def user(message, files, history, session_id, save_history):
        if files:
            for file_path in files:
                try:
                    file_extension = os.path.splitext(file_path)[1].lower()
                    file_content = ""

                    if file_extension == ".pdf":
                        with pdfplumber.open(file_path) as pdf:
                            for page in pdf.pages:
                                file_content += page.extract_text()
                    elif file_extension == ".docx":
                        doc = docx.Document(file_path)
                        for paragraph in doc.paragraphs:
                            file_content += paragraph.text + "\n"
                    elif file_extension == ".xlsx":
                        workbook = load_workbook(file_path)
                        for sheet in workbook.sheetnames:
                            worksheet = workbook[sheet]
                            for row in worksheet.iter_rows():
                                row_values = [str(cell.value) for cell in row]
                                file_content += ", ".join(row_values) + "\n"
                    else:
                        message += f"\nUnsupported file type: {file_extension}"
                        continue

                    message += f"\nFile content from {file_path}:\n{file_content}"

                except Exception as e:
                    message += f"\nError processing {file_path}: {str(e)}"

        if save_history:
            session = session_manager.load_session(session_id)
            session["history"].append({
                "timestamp": datetime.now().isoformat(),
                "type": "user",
                "content": message
            })
            session_manager.save_session(session_id,session)
            return "", history + [[message, None]]


    def bot(history, session_id):
        if history and history[-1][1] is None:
            message = history[-1][0]
            for response in respond(message, history[:-1], session_id):
                history[-1][1] = response
                yield history

    with gr.Row():
        msg = gr.Textbox(label="Message")
        file_upload = gr.File(file_types=[".pdf", ".docx", ".xlsx"], file_count="multiple")

    msg.submit(user, [msg, file_upload, chatbot, session_id, save_history], [msg, chatbot]).then(
        bot, [chatbot, session_id], [chatbot]
    )
    new_session.click(on_new_session, None, [session_id, chatbot])

if __name__ == "__main__":
    demo.launch(share=True)
